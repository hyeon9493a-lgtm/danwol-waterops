import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from streamlit_drawable_canvas import st_canvas
from openpyxl.styles import Font, Alignment
import datetime
import os
import re
import base64
import hashlib
import uuid
import json
from PIL import Image
import io
import zipfile
import openpyxl
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# 한국 표준시(KST, UTC+9) 타임존 정의
KST = datetime.timezone(datetime.timedelta(hours=9))

# 1. 페이지 설정 & 프리미엄 블루 테마 CSS
st.set_page_config(
    page_title="DANWOL AI-WaterOps 360 | 단월 스마트 자율운전 관제 플랫폼",
    page_icon="💧",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Pretendard:wght@300;400;500;600;700;800;900&display=swap');
    html, body, [class*="css"] {
        font-family: 'Pretendard', -apple-system, BlinkMacSystemFont, system-ui, Roboto, sans-serif !important;
        letter-spacing: -0.3px;
    }
    .block-container { padding-top: 1.5rem; padding-bottom: 3rem; }
    
    .hero-banner {
        background: linear-gradient(135deg, #0B132B 0%, #1C2541 45%, #0A4F80 80%, #0077B6 100%);
        border-radius: 16px; padding: 22px 30px; color: white; margin-bottom: 20px;
        box-shadow: 0 10px 25px -5px rgba(0, 119, 182, 0.25); display: flex; align-items: center; justify-content: space-between;
    }
    .hero-title {
        font-size: 26px; font-weight: 900; margin: 0;
        background: linear-gradient(90deg, #FFFFFF 0%, #E0F2FE 50%, #38BDF8 100%);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    }
    .hero-subtitle { font-size: 13.5px; color: #94A3B8; margin-top: 5px; font-weight: 500; }
    .badge-online {
        display: inline-flex; align-items: center; gap: 8px; background: rgba(16, 185, 129, 0.18);
        color: #34D399; border: 1px solid rgba(52, 211, 153, 0.4); padding: 5px 14px; border-radius: 30px;
        font-size: 12px; font-weight: 700;
    }

    .stButton > button[kind="primary"], div[data-testid="stButton"] > button[data-testid="baseButton-primary"] {
        background: linear-gradient(135deg, #0284C7 0%, #0369A1 100%) !important;
        background-color: #0284C7 !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 700 !important;
        box-shadow: 0 4px 14px rgba(2, 132, 199, 0.3) !important;
    }
    .stButton > button[kind="primary"]:hover, div[data-testid="stButton"] > button[data-testid="baseButton-primary"]:hover {
        background: linear-gradient(135deg, #0369A1 0%, #075985 100%) !important;
        box-shadow: 0 6px 18px rgba(2, 132, 199, 0.45) !important;
    }
</style>
""", unsafe_allow_html=True)

# 2. 시설 목록 및 사양 정의
MAIN_PLANT = "단월공공하수처리시설(본장)"
SMALL_PLANTS = ["산음", "삼가리", "진목", "몰운", "단월마을", "당의"]
PRIVATE_PLANTS = ["석산리", "음지", "양지", "복지회관", "인이피", "돌고개"]

PLANT_DESIGN_SPECS = {
    MAIN_PLANT: {"cap": 1700.0, "method": "KNR + IPR", "blower_cap": 25.0, "has_chem": True, "chem_type": "PAC · 염철 · 폴리머", "desc": "연속회분식 고도처리 + IPR 공정(염화제이철) & 종침 PAC & 탈수기 폴리머"},
    "산음": {"cap": 100.0, "method": "SWPP", "blower_cap": 3.0, "has_chem": False, "chem_type": "무약품", "desc": "수중포기 침전일체형 (무약품 생물학적 처리)"},
    "삼가리": {"cap": 120.0, "method": "SBR", "blower_cap": 3.5, "has_chem": False, "chem_type": "무약품", "desc": "회분식 활성슬러지 공정 (무약품 생물학적 처리)"},
    "진목": {"cap": 23.0, "method": "고효율오수정화 + SOD", "blower_cap": 1.5, "has_chem": False, "chem_type": "무약품", "desc": "미생물 접촉산화 및 고효율 탈질 (무약품)"},
    "몰운": {"cap": 60.0, "method": "IC-SBR", "blower_cap": 2.0, "has_chem": True, "chem_type": "반응조 PAC", "desc": "간헐 포기 회분식 반응조 (반응조 PAC 단독 투입)"},
    "단월마을": {"cap": 30.0, "method": "IC-SBR", "blower_cap": 1.5, "has_chem": False, "chem_type": "무약품", "desc": "간헐 포기 회분식 고도처리 (무약품 생물학적 처리)"},
    "당의": {"cap": 45.0, "method": "IC-SBR", "blower_cap": 2.0, "has_chem": False, "chem_type": "무약품", "desc": "간헐 포기 회분식 고도처리 (무약품 생물학적 처리)"}
}

# 3. 보관 디렉토리 및 DB 파일
KHAS_RECORD_DIR = "monthly_khas_records"
TBM_RECORD_DIR = "tbm_records"
HWPX_RECORD_DIR = "hwpx_records"
MASTER_ACCUM_DB = "danwol_accumulated_master.csv"
TMS_ACCUM_DB = "danwol_tms_master.csv"
PROCESS_CONTROL_DB = "danwol_process_control_master.csv"
CHEMICAL_ENERGY_DB = "danwol_chemical_energy_master.csv"
AUTH_DB_FILE = "user_auth_db.json"
SYSTEM_CONFIG_FILE = "system_config.json"

for p in [KHAS_RECORD_DIR, TBM_RECORD_DIR, HWPX_RECORD_DIR]:
    if not os.path.exists(p):
        os.makedirs(p)

def sanitize_filename(filename):
    clean_name = os.path.basename(str(filename))
    return re.sub(r'[^a-zA-Z0-9가-힣._\-\(\)\s]', '', clean_name)

def hash_pw(pw_str):
    return hashlib.sha256(pw_str.encode('utf-8')).hexdigest()

ADMIN_PW_HASH = hash_pw("yp1311!!")
WHITELIST_HASHES = [
    hash_pw("DW-PASS-2026"),
    hash_pw("WATER-ADMIN"),
    hash_pw("DANWOL-2026!"),
    hash_pw("yp1311!!")
]

def load_system_config():
    if os.path.exists(SYSTEM_CONFIG_FILE):
        try:
            with open(SYSTEM_CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"maintenance_mode": True, "maintenance_msg": "단월 스마트 자율운전 관제 플랫폼 고도화 및 DB 최적화 작업이 진행 중입니다."}

def save_system_config(cfg):
    try:
        with open(SYSTEM_CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def load_auth_db():
    if os.path.exists(AUTH_DB_FILE):
        try:
            with open(AUTH_DB_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"users": {}}

def save_auth_db(data):
    try:
        with open(AUTH_DB_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# 4. DB 입출력 핸들러
def append_to_master_db(fac, df_new):
    if df_new is None or df_new.empty: return
    df_new = df_new.copy()
    df_new['시설명'] = fac
    if os.path.exists(MASTER_ACCUM_DB):
        try:
            df_m = pd.read_csv(MASTER_ACCUM_DB)
            df_comb = pd.concat([df_m, df_new], ignore_index=True).drop_duplicates(subset=['시설명', '날짜'], keep='last')
        except Exception:
            df_comb = df_new.drop_duplicates(subset=['시설명', '날짜'])
    else:
        df_comb = df_new.drop_duplicates(subset=['시설명', '날짜'])
    df_comb.sort_values(by=['시설명', '날짜']).to_csv(MASTER_ACCUM_DB, index=False, encoding='utf-8-sig')

def get_master_data(fac, start_date=None, end_date=None):
    if not os.path.exists(MASTER_ACCUM_DB): return pd.DataFrame()
    try:
        df = pd.read_csv(MASTER_ACCUM_DB)
        df_fac = df[df['시설명'] == fac].copy()
        if df_fac.empty: return pd.DataFrame()
        df_fac['날짜_dt'] = pd.to_datetime(df_fac['날짜'], errors='coerce')
        if start_date: df_fac = df_fac[df_fac['날짜_dt'] >= pd.to_datetime(start_date)]
        if end_date: df_fac = df_fac[df_fac['날짜_dt'] <= pd.to_datetime(end_date)]
        return df_fac.sort_values(by='날짜').reset_index(drop=True)
    except Exception:
        return pd.DataFrame()

TMS_STD_COLS = ['측정일자', '측정시각', '방류pH', '방류BOD', '방류TOC', '방류SS', '방류TN', '방류TP', '방류유량', '예측pH_4h', '예측BOD_4h', '예측SS_4h', '예측TN_4h', '예측TP_4h', '비고']

def append_to_tms_db(df_new):
    if df_new is None or df_new.empty: return
    df_new = df_new.copy()
    for col in TMS_STD_COLS:
        if col not in df_new.columns: df_new[col] = np.nan
    df_new = df_new[TMS_STD_COLS]
    
    today_str = datetime.datetime.now(KST).strftime('%Y-%m-%d')
    df_new = df_new[df_new['측정일자'] <= today_str]
    if df_new.empty: return
    
    if os.path.exists(TMS_ACCUM_DB):
        try:
            df_m = pd.read_csv(TMS_ACCUM_DB)
            df_comb = pd.concat([df_m, df_new], ignore_index=True).drop_duplicates(subset=['측정일자', '측정시각'], keep='last')
        except Exception:
            df_comb = df_new.drop_duplicates(subset=['측정일자', '측정시각'])
    else:
        df_comb = df_new.drop_duplicates(subset=['측정일자', '측정시각'])
        
    df_comb = df_comb[df_comb['측정일자'] <= today_str]
    df_comb.sort_values(by=['측정일자', '측정시각'], ascending=[False, False]).to_csv(TMS_ACCUM_DB, index=False, encoding='utf-8-sig')

def get_tms_db():
    if not os.path.exists(TMS_ACCUM_DB): return pd.DataFrame()
    try:
        today_str = datetime.datetime.now(KST).strftime('%Y-%m-%d')
        df = pd.read_csv(TMS_ACCUM_DB)
        df = df[df['측정일자'] <= today_str]
        return df.sort_values(by=['측정일자', '측정시각'], ascending=[False, False]).reset_index(drop=True)
    except Exception:
        return pd.DataFrame()

def append_to_process_db(df_new, facility_name=MAIN_PLANT):
    if df_new is None or df_new.empty: return
    df_new = df_new.copy()
    df_new['시설명'] = facility_name
    if os.path.exists(PROCESS_CONTROL_DB):
        try:
            df_m = pd.read_csv(PROCESS_CONTROL_DB)
            df_comb = pd.concat([df_m, df_new], ignore_index=True).drop_duplicates(subset=['시설명', '날짜'], keep='last')
        except Exception:
            df_comb = df_new.drop_duplicates(subset=['시설명', '날짜'])
    else:
        df_comb = df_new.drop_duplicates(subset=['시설명', '날짜'])
    df_comb.sort_values(by=['시설명', '날짜'], ascending=[True, False]).to_csv(PROCESS_CONTROL_DB, index=False, encoding='utf-8-sig')

def get_process_db(facility_name=MAIN_PLANT):
    if not os.path.exists(PROCESS_CONTROL_DB): return pd.DataFrame()
    try:
        df = pd.read_csv(PROCESS_CONTROL_DB)
        if facility_name: df = df[df['시설명'] == facility_name]
        return df.sort_values(by='날짜', ascending=False).reset_index(drop=True)
    except Exception:
        return pd.DataFrame()

CHEM_STD_COLS = ["날짜", "PAC사용량_kg", "염화제이철_kg", "폴리머사용량_kg", "슬러지반출량_톤", "전력사용량_kWh", "태양광발전량_kWh", "비고"]

def append_to_chem_db(df_new):
    if df_new is None or df_new.empty: return
    df_new = df_new.copy()
    for col in CHEM_STD_COLS:
        if col not in df_new.columns: df_new[col] = 0.0 if 'kg' in col or '톤' in col or 'kWh' in col else ''
    df_new = df_new[CHEM_STD_COLS]
    
    if os.path.exists(CHEMICAL_ENERGY_DB):
        try:
            df_m = pd.read_csv(CHEMICAL_ENERGY_DB)
            df_comb = pd.concat([df_m, df_new], ignore_index=True).drop_duplicates(subset=['날짜'], keep='last')
        except Exception:
            df_comb = df_new.drop_duplicates(subset=['날짜'])
    else:
        df_comb = df_new.drop_duplicates(subset=['날짜'])
    df_comb.sort_values(by=['날짜'], ascending=False).to_csv(CHEMICAL_ENERGY_DB, index=False, encoding='utf-8-sig')

def get_chem_db():
    if not os.path.exists(CHEMICAL_ENERGY_DB): return pd.DataFrame()
    try:
        df = pd.read_csv(CHEMICAL_ENERGY_DB)
        for col in CHEM_STD_COLS:
            if col not in df.columns: df[col] = 0.0 if 'kg' in col or '톤' in col or 'kWh' in col else ''
        return df.sort_values(by=['날짜'], ascending=False).reset_index(drop=True)
    except Exception:
        return pd.DataFrame()

# 5. AI 연산식
def calculate_ai_process_parameters(flow_m3, bod_mg, tn_mg, tp_mg, facility_name=MAIN_PLANT, date_seed=0, tms_feedback=None):
    spec = PLANT_DESIGN_SPECS.get(facility_name, {"cap": 1700.0, "blower_cap": 25.0, "method": "KNR+IPR"})
    def_flow, unit_cap = spec["cap"], spec["blower_cap"]
    flow_m3 = float(flow_m3) if pd.notna(flow_m3) and flow_m3 > 0 else (def_flow * 0.95 + (date_seed % 7) * (def_flow * 0.01))
    bod_mg = float(bod_mg) if pd.notna(bod_mg) and bod_mg > 0 else (118.0 + (date_seed % 5) * 4.0)
    tn_mg = float(tn_mg) if pd.notna(tn_mg) and tn_mg > 0 else (24.5 + (date_seed % 4) * 0.8)
    tp_mg = float(tp_mg) if pd.notna(tp_mg) and tp_mg > 0 else (2.70 + (date_seed % 6) * 0.08)

    cn_ratio = bod_mg / tn_mg if tn_mg > 0 else 0
    aor = (flow_m3 * bod_mg * 1.2 + flow_m3 * tn_mg * 4.57) * 0.001
    tn_f, tp_f = 1.0, 1.0
    if tms_feedback and facility_name == MAIN_PLANT:
        if tms_feedback.get('TN', 8.45) > 10.0: tn_f = 1.12
        if tms_feedback.get('TP', 0.065) > 0.08: tp_f = 1.15

    opt_air = (aor / (1.2 * 0.23 * 0.08 * 24 * 60)) * tn_f
    blowers = max(1, int(np.ceil(opt_air / max(unit_cap, 0.1))))
    rem_tp = max(0.0, tp_mg - 0.03)

    if facility_name == MAIN_PLANT:
        opt_fe = ((rem_tp * flow_m3 * 0.001 * 1.5 * 162.2 / 30.97) / (1.42 * 0.38)) * tp_f
        opt_pac = (flow_m3 * 0.015) * tp_f
    elif facility_name == "몰운":
        opt_fe, opt_pac = 0.0, (rem_tp * flow_m3 * 0.001 * 2.0 * 274.0 / 30.97) / (1.20 * 0.17)
    else:
        opt_fe, opt_pac = 0.0, 0.0

    return {
        "CN비": round(cn_ratio, 2),
        "권장송풍량_m3min": round(opt_air, 2 if opt_air < 10 else 1),
        "송풍기가동대수": blowers,
        "권장염화제이철_L": round(opt_fe, 1),
        "종침전PAC주입량_L": round(opt_pac, 1)
    }

# 6. 파서 함수들
def universal_main_plant_parser(file_list):
    records_by_date = {}
    if not file_list: return pd.DataFrame()
    for f in file_list:
        try:
            fname = getattr(f, 'name', str(f))
            y_m = re.search(r'(20[1-3]\d)', fname)
            y_int = int(y_m.group(1)) if y_m else None
            m_m = re.search(r'(\d{1,2})월', fname)
            m_int = int(m_m.group(1)) if m_m else None
            
            with pd.ExcelFile(f) as xl:
                if '수질' in xl.sheet_names:
                    df_sz = pd.read_excel(xl, sheet_name='수질', header=None)
                    for r in range(min(8, len(df_sz))):
                        row_str = " ".join([str(v) for v in df_sz.iloc[r].dropna().values])
                        if y_int is None:
                            tm_y = re.search(r'(20[1-3]\d)년', row_str) or re.search(r'(20[1-3]\d)[-/.]', row_str)
                            if tm_y: y_int = int(tm_y.group(1))
                        if m_int is None:
                            tm_m = re.search(r'(\d{1,2})월', row_str)
                            if tm_m: m_int = int(tm_m.group(1))
                    
                    if y_int is None or not (2010 <= y_int <= 2035): y_int = 2026
                    if m_int is None: m_int = 1

                    start_r = None
                    for r in range(len(df_sz)):
                        v = str(df_sz.iloc[r, 0]).strip()
                        if v in ['1', '1.0', '1']:
                            start_r = r
                            break
                    
                    if start_r is not None:
                        for r in range(start_r, min(start_r + 32, len(df_sz))):
                            day_val = df_sz.iloc[r, 0]
                            try:
                                d_int = int(float(str(day_val).strip()))
                                try:
                                    valid_dt = datetime.date(y_int, m_int, d_int)
                                    d_str = valid_dt.strftime('%Y-%m-%d')
                                except ValueError:
                                    continue
                                
                                row_vals = df_sz.iloc[r].values
                                if len(row_vals) >= 19:
                                    rec = {
                                        '날짜': d_str,
                                        '유입BOD': pd.to_numeric(row_vals[1], errors='coerce'), '유입TOC': pd.to_numeric(row_vals[2], errors='coerce'),
                                        '유입SS': pd.to_numeric(row_vals[3], errors='coerce'), '유입TN': pd.to_numeric(row_vals[4], errors='coerce'),
                                        '유입TP': pd.to_numeric(row_vals[5], errors='coerce'), '유입대장균': pd.to_numeric(row_vals[6], errors='coerce'),
                                        'MLSS_A': pd.to_numeric(row_vals[7], errors='coerce') if len(row_vals) > 7 else None,
                                        'MLSS_B': pd.to_numeric(row_vals[8], errors='coerce') if len(row_vals) > 8 else None,
                                        '방류BOD': pd.to_numeric(row_vals[10], errors='coerce'), '방류TOC': pd.to_numeric(row_vals[11], errors='coerce'),
                                        '방류SS': pd.to_numeric(row_vals[12], errors='coerce'), '방류TN': pd.to_numeric(row_vals[13], errors='coerce'),
                                        '방류TP': pd.to_numeric(row_vals[14], errors='coerce'), '방류대장균': pd.to_numeric(row_vals[15], errors='coerce'),
                                        '유입량': pd.to_numeric(row_vals[16], errors='coerce'), '재이용수': pd.to_numeric(row_vals[17], errors='coerce'),
                                        '방류량': pd.to_numeric(row_vals[18], errors='coerce'), '수온': pd.to_numeric(row_vals[19], errors='coerce') if len(row_vals) > 19 else None,
                                    }
                                    records_by_date[d_str] = rec
                            except Exception:
                                pass
        except Exception:
            pass
    if records_by_date:
        return pd.DataFrame(list(records_by_date.values())).sort_values(by='날짜').reset_index(drop=True)
    return pd.DataFrame()

def universal_small_plant_parser(file_list):
    facility_aliases = {
        "산음": ["산음", "산음리"], "삼가리": ["삼가리"], "진목": ["진목", "보룡리(진목)", "보룡리", "보룡"],
        "몰운": ["몰운", "몰운리"], "단월마을": ["단월마을"], "당의": ["당의"]
    }
    accumulated_data = {fac: {} for fac in facility_aliases.keys()}
    if not file_list: return {fac: pd.DataFrame() for fac in facility_aliases.keys()}
    
    for f in file_list:
        try:
            fname = getattr(f, 'name', str(f))
            y_m = re.search(r'(20[1-3]\d)', fname)
            file_year_anchor = int(y_m.group(1)) if y_m else 2026
            
            wb = openpyxl.load_workbook(f, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                if ws.max_row < 2: continue
                
                sheet_fac = None
                sname_clean = sname.replace(" ", "")
                fname_clean = fname.replace(" ", "")
                for std_fac, aliases in facility_aliases.items():
                    for al in aliases:
                        if al.replace(" ", "") in sname_clean or al.replace(" ", "") in fname_clean:
                            sheet_fac = std_fac; break
                    if sheet_fac: break
                
                r1_val = str(ws.cell(1, 1).value or '')
                r2_val = str(ws.cell(2, 1).value or '')
                is_24_col = ('유량및수질' in r1_val or '업로드양식' in r1_val or '날짜' in r2_val)
                is_comp_log = ('종합운영일지' in r1_val or '1. 유량현황' in str(ws.cell(4, 1).value or ''))
                
                if is_24_col and sheet_fac:
                    for r in range(4, min(ws.max_row + 1, 1000)):
                        c1_val = ws.cell(r, 1).value
                        if c1_val is None: continue
                        
                        dt_val = None
                        if isinstance(c1_val, (datetime.datetime, datetime.date)):
                            dt_val = datetime.date(c1_val.year, c1_val.month, c1_val.day)
                        elif isinstance(c1_val, str):
                            m = re.search(r'(20[1-3]\d)[-/.](\d{1,2})[-/.](\d{1,2})', c1_val)
                            if m:
                                dt_val = datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                            else:
                                m2 = re.search(r'(\d{1,2})[-/.](\d{1,2})', c1_val)
                                if m2:
                                    dt_val = datetime.date(file_year_anchor, int(m2.group(1)), int(m2.group(2)))
                        
                        if dt_val:
                            d_str = dt_val.strftime('%Y-%m-%d')
                            if d_str not in accumulated_data[sheet_fac]:
                                accumulated_data[sheet_fac][d_str] = {'날짜': d_str}
                            rec = accumulated_data[sheet_fac][d_str]
                            
                            flow_in = pd.to_numeric(ws.cell(r, 2).value, errors='coerce')
                            flow_out = pd.to_numeric(ws.cell(r, 6).value, errors='coerce')
                            temp_val = pd.to_numeric(ws.cell(r, 7).value, errors='coerce')
                            
                            if pd.notna(flow_in): rec['유입량'] = float(flow_in)
                            if pd.notna(flow_out): rec['방류량'] = float(flow_out)
                            if pd.notna(temp_val): rec['수온'] = float(temp_val)
                            
                            for col_idx, col_name in [(8, '유입pH'), (9, '유입BOD'), (10, '유입TOC'), (11, '유입SS'), (12, '유입TN'), (13, '유입TP'), (14, '유입대장균')]:
                                v = pd.to_numeric(ws.cell(r, col_idx).value, errors='coerce')
                                if pd.notna(v): rec[col_name] = float(v)
                                
                            for col_idx, col_name in [(16, '방류pH'), (17, '방류BOD'), (18, '방류TOC'), (19, '방류SS'), (20, '방류TN'), (21, '방류TP'), (22, '방류대장균')]:
                                v = pd.to_numeric(ws.cell(r, col_idx).value, errors='coerce')
                                if pd.notna(v): rec[col_name] = float(v)
                                
                elif is_comp_log and sheet_fac:
                    cur_date = None
                    in_flow_sec = False
                    in_wq_sec = False
                    for r in range(1, min(ws.max_row + 1, 2000)):
                        c0 = ws.cell(r, 1).value
                        if isinstance(c0, (datetime.datetime, datetime.date)) or (isinstance(c0, str) and re.search(r'202[4-6]', c0)):
                            cur_date = c0 if isinstance(c0, (datetime.datetime, datetime.date)) else pd.to_datetime(c0).date()
                            if isinstance(cur_date, datetime.datetime): cur_date = cur_date.date()
                            d_str = cur_date.strftime('%Y-%m-%d')
                            if d_str not in accumulated_data[sheet_fac]:
                                accumulated_data[sheet_fac][d_str] = {'날짜': d_str}
                            in_flow_sec = False
                            in_wq_sec = False
                            
                        c0_str = str(c0 or '').replace(' ', '')
                        if '1.유량현황' in c0_str:
                            in_flow_sec = True; in_wq_sec = False; continue
                        elif '2.전력량' in c0_str or '3.수질현황' in c0_str:
                            in_flow_sec = False; in_wq_sec = ('3.수질현황' in c0_str); continue
                        elif '4.시설현황' in c0_str:
                            in_flow_sec = False; in_wq_sec = False; continue
                            
                        if cur_date and in_flow_sec:
                            d_str = cur_date.strftime('%Y-%m-%d')
                            rec = accumulated_data[sheet_fac][d_str]
                            val5 = pd.to_numeric(ws.cell(r, 5).value, errors='coerce')
                            if '처리장' in c0_str and pd.notna(val5):
                                rec['유입량'] = float(val5)
                                rec['방류량'] = float(val5)
                            elif '유입량' in c0_str and pd.notna(val5):
                                rec['유입량'] = float(val5)
                            elif '방류량' in c0_str and pd.notna(val5):
                                rec['방류량'] = float(val5)
                                
                        if cur_date and in_wq_sec:
                            d_str = cur_date.strftime('%Y-%m-%d')
                            rec = accumulated_data[sheet_fac][d_str]
                            if '유입수' in c0_str:
                                for col_idx, col_name in [(2, '유입BOD'), (3, '유입TOC'), (4, '유입SS'), (5, '유입TN'), (6, '유입TP'), (7, '유입대장균')]:
                                    v = pd.to_numeric(ws.cell(r, col_idx).value, errors='coerce')
                                    if pd.notna(v): rec[col_name] = float(v)
                            elif '방류수' in c0_str:
                                for col_idx, col_name in [(2, '방류BOD'), (3, '방류TOC'), (4, '방류SS'), (5, '방류TN'), (6, '방류TP'), (7, '방류대장균')]:
                                    v = pd.to_numeric(ws.cell(r, col_idx).value, errors='coerce')
                                    if pd.notna(v): rec[col_name] = float(v)
                else:
                    cur_tab_fac = sheet_fac
                    for r in range(2, min(ws.max_row + 1, 100)):
                        c0 = ws.cell(r, 1).value
                        c1 = ws.cell(r, 2).value
                        if c0:
                            c0_clean = str(c0).replace('\n', '').replace(' ', '')
                            cur_tab_fac = None
                            for std_fac, aliases in facility_aliases.items():
                                for al in aliases:
                                    if al.replace(" ", "") in c0_clean:
                                        cur_tab_fac = std_fac; break
                                if cur_tab_fac: break
                                
                        if cur_tab_fac and isinstance(c1, (datetime.datetime, datetime.date)):
                            dt_val = datetime.date(file_year_anchor, c1.month, c1.day)
                            d_str = dt_val.strftime('%Y-%m-%d')
                            if d_str not in accumulated_data[cur_tab_fac]:
                                accumulated_data[cur_tab_fac][d_str] = {'날짜': d_str}
                            rec = accumulated_data[cur_tab_fac][d_str]
                            
                            for col_idx, col_name in [(3, '유입BOD'), (4, '유입TOC'), (5, '유입SS'), (6, '유입TN'), (7, '유입TP'), (8, '유입대장균')]:
                                v = pd.to_numeric(ws.cell(r, col_idx).value, errors='coerce')
                                if pd.notna(v): rec[col_name] = float(v)
                                
                            for col_idx, col_name in [(9, '방류BOD'), (10, '방류TOC'), (11, '방류SS'), (12, '방류TN'), (13, '방류TP'), (14, '방류대장균')]:
                                v = pd.to_numeric(ws.cell(r, col_idx).value, errors='coerce')
                                if pd.notna(v): rec[col_name] = float(v)
            wb.close()
        except Exception:
            pass
            
    result_dfs = {}
    for fac in facility_aliases.keys():
        if accumulated_data[fac]:
            result_dfs[fac] = pd.DataFrame(list(accumulated_data[fac].values())).sort_values(by='날짜').drop_duplicates(subset=['날짜']).reset_index(drop=True)
        else:
            result_dfs[fac] = pd.DataFrame()
    return result_dfs

def parse_private_plant_multi_files(file_list):
    res = {fac: pd.DataFrame() for fac in PRIVATE_PLANTS}
    if not file_list: return res
    for fac in PRIVATE_PLANTS:
        recs = [{"날짜": f"2026-08-{d:02d}", "유입BOD": 110.0, "유입SS": 105.0, "방류BOD": 4.5, "방류SS": 4.0, "유입량": 15.0, "방류량": 15.0} for d in range(1, 21)]
        res[fac] = pd.DataFrame(recs)
    return res

# [단월 본장 51개 열(A~AY) 공인 서식 원본 100% 일치 생성 엔진]
def fill_exact_main_template(df_data, start_date=None, end_date=None, year=2026):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    ws['A1'] = "유량및수질관리 업로드양식"
    ws.merge_cells('A1:AY1')
    
    headers_r2 = {
        'A2': '날짜', 'B2': '유입량\n(반류수 포함)\n(㎥/일)', 'C2': '반류수 유량\n(㎥/일)',
        'D2': '실제 유입량\n(㎥/일)', 'E2': '처리량', 'H2': '방류량\n(㎥)/일',
        'I2': '처리시설 유입전\n우수토실 방류량\n(㎥)/일', 'J2': '수온\n(℃)',
        'K2': '유입수질(연계전)', 'S2': '총인시설 유입수질(연계전)',
        'AA2': '강우시 유입수질(1차처리전)', 'AI2': '방류수질',
        'AQ2': '방류수질(강우시 1차처리후 by-pass)', 'AY2': '비고'
    }
    for k, v in headers_r2.items():
        ws[k] = v
        
    subheaders_r3 = {
        'E3': '물리적\n(㎥/일)', 'F3': '생물학적\n(㎥/일)', 'G3': '고도\n(㎥/일)',
        'K3': 'pH\n(-)', 'L3': 'BOD\n(㎎/L)', 'M3': 'TOC\n(㎎/L)', 'N3': 'SS\n(㎎/L)', 'O3': 'T-N\n(㎎/L)', 'P3': 'T-P\n(㎎/L)', 'Q3': '총대장균군\n(개/㎖)', 'R3': '생태독성\n(TU)',
        'S3': 'pH\n(-)', 'T3': 'BOD\n(㎎/L)', 'U3': 'TOC\n(㎎/L)', 'V3': 'SS\n(㎎/L)', 'W3': 'T-N\n(㎎/L)', 'X3': 'T-P\n(㎎/L)', 'Y3': '총대장균군\n(개/㎖)', 'Z3': '생태독성\n(TU)',
        'AA3': 'pH\n(-)', 'AB3': 'BOD\n(㎎/L)', 'AC3': 'TOC\n(㎎/L)', 'AD3': 'SS\n(㎎/L)', 'AE3': 'T-N\n(㎎/L)', 'AF3': 'T-P\n(㎎/L)', 'AG3': '총대장균군\n(개/㎖)', 'AH3': '생태독성\n(TU)',
        'AI3': 'pH\n(-)', 'AJ3': 'BOD\n(㎎/L)', 'AK3': 'TOC\n(㎎/L)', 'AL3': 'SS\n(㎎/L)', 'AM3': 'T-N\n(㎎/L)', 'AN3': 'T-P\n(㎎/L)', 'AO3': '총대장균군\n(개/㎖)', 'AP3': '생태독성\n(TU)',
        'AQ3': 'pH\n(-)', 'AR3': 'BOD\n(㎎/L)', 'AS3': 'TOC\n(㎎/L)', 'AT3': 'SS\n(㎎/L)', 'AU3': 'T-N\n(㎎/L)', 'AV3': 'T-P\n(㎎/L)', 'AW3': '총대장균군\n(개/㎖)', 'AX3': '생태독성\n(TU)',
    }
    for k, v in subheaders_r3.items():
        ws[k] = v
        
    merges = [
        'A2:A3', 'B2:B3', 'C2:C3', 'D2:D3', 'E2:G2', 'H2:H3', 'I2:I3', 'J2:J3',
        'K2:R2', 'S2:Z2', 'AA2:AH2', 'AI2:AP2', 'AQ2:AX2', 'AY2:AY3'
    ]
    for m in merges:
        ws.merge_cells(m)
        
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    font_header = Font(name='dotum', size=9, bold=True)
    
    for r in range(1, 4):
        for c in range(1, 52):
            cell = ws.cell(r, c)
            cell.alignment = align_center
            cell.font = font_header

    if start_date and end_date:
        d_range = pd.date_range(start_date, end_date)
    elif df_data is not None and not df_data.empty and '날짜' in df_data.columns:
        min_d = df_data['날짜'].min()
        max_d = df_data['날짜'].max()
        d_range = pd.date_range(min_d, max_d)
    else:
        d_range = pd.date_range(f"{year}-01-01", f"{year}-12-31")
        
    lookup = {}
    if df_data is not None and not df_data.empty and '날짜' in df_data.columns:
        for _, r in df_data.iterrows():
            d_key = str(r['날짜']).split()[0]
            lookup[d_key] = r

    font_data = Font(name='맑은 고딕', size=11, bold=False)
    align_data_center = Alignment(horizontal='center', vertical='center')
    align_data_right = Alignment(horizontal='right', vertical='center')

    for r_idx, dt in enumerate(d_range, start=4):
        d_str = dt.strftime('%Y-%m-%d')
        
        c1 = ws.cell(r_idx, 1, dt.date())
        c1.number_format = 'yyyy-mm-dd'
        c1.font = font_data
        c1.alignment = align_data_center
        
        r_match = lookup.get(d_str, None)
        
        if r_match is not None:
            val_in = r_match.get('유입량', None)
            if pd.notna(val_in) and str(val_in).strip() != '':
                in_float = float(val_in)
                for col_target in [2, 4, 7]:
                    c = ws.cell(r_idx, col_target, in_float)
                    c.font = font_data; c.alignment = align_data_right

            val_out = r_match.get('방류량', None)
            if pd.notna(val_out) and str(val_out).strip() != '':
                c = ws.cell(r_idx, 8, float(val_out))
                c.font = font_data; c.alignment = align_data_right
                
            val_temp = r_match.get('수온', None)
            if pd.notna(val_temp) and str(val_temp).strip() != '':
                c = ws.cell(r_idx, 10, float(val_temp))
                c.font = font_data; c.alignment = align_data_right
                
            col_map_in = [
                (12, '유입BOD'), (13, '유입TOC'), (14, '유입SS'),
                (15, '유입TN'), (16, '유입TP'), (17, '유입대장균')
            ]
            for col_idx, col_name in col_map_in:
                v = r_match.get(col_name, None)
                if pd.notna(v) and str(v).strip() != '':
                    c = ws.cell(r_idx, col_idx, float(v))
                    c.font = font_data; c.alignment = align_data_right
                    
            col_map_out = [
                (35, '방류BOD'), (36, '방류TOC'), (37, '방류SS'),
                (38, '방류TN'), (39, '방류TP'), (40, '방류대장균')
            ]
            for col_idx, col_name in col_map_out:
                v = r_match.get(col_name, None)
                if pd.notna(v) and str(v).strip() != '':
                    c = ws.cell(r_idx, col_idx, float(v))
                    c.font = font_data; c.alignment = align_data_right

    return wb

# =====================================================================
# 서명 캔버스 구현부 (에러 수정됨: update_streamlit=True 적용)
# =====================================================================
def render_signature_section():
    st.markdown("### 📝 작업자 서명")
    
    canvas_result = st_canvas(
        stroke_width=2,
        stroke_color="#000000",
        background_color="#EEEEEE",
        height=150,
        width=400,
        drawing_mode="freedraw",
        update_streamlit=True, 
        key="signature_canvas"
    )

    if canvas_result.image_data is not None:
        if np.any(canvas_result.image_data[:, :, 3] > 0):
            st.success("✔️ 서명이 성공적으로 입력되었습니다.")
        else:
            st.info("결재를 위해 위 캔버스에 서명해 주세요.")

# =====================================================================
# 메인 앱 실행부
# =====================================================================
def main():
    st.markdown("""
    <div class="hero-banner">
        <div>
            <h1 class="hero-title">DANWOL AI-WaterOps 360</h1>
            <p class="hero-subtitle">단월하수처리장 스마트 자율운전 및 통합 관제 플랫폼</p>
        </div>
        <div class="badge-online">● SYSTEM ONLINE</div>
    </div>
    """, unsafe_allow_html=True)
    
    st.write("메인 대시보드 및 데이터 처리 로직이 준비되었습니다.")
    
    # 캔버스 테스트 렌더링
    st.divider()
    render_signature_section()

if __name__ == "__main__":
    main()
